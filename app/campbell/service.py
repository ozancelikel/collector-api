import csv
import traceback
from datetime import datetime

from fastapi import Depends
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook

from app.campbell.schemas import CampbellMessage
from app.campbell.scraper_scripts.loginKonect import init_driver, login_to_konect
from app.campbell.scraper_scripts.task_query import open_task_query, modify_query_date
from app.campbell.scraper_scripts.download_file import click_download_button, move_and_rename_file
from app.config import settings
from app.entities.file_type import FileType
from app.logs.config_scraper_logs import scraper_logger
from app.db.session import get_db
from app.models.models import CampbellSensors
from app.logs.config_server_logs import server_logger
from app.db.session import AsyncSessionLocal


async def run_campbell_scraper(hourly: bool=True):
    """Programme principal qui gère la connexion et les tâches d'automatisation."""
    driver = init_driver(visibility=False)

    try:
        # **Se connecter à KonectGDS**
        driver = login_to_konect(
            driver=driver,
            username=settings.KONECTGDS_USERNAME,
            password=settings.KONECTGDS_PASSWORD
        )

        # **Ouvrir Table Query et modifier la date**
        open_task_query(driver, hourly)
        modify_query_date(driver, hourly)

        if settings.SCRAPER_DOWNLOAD_FILE_TYPE == '.csv':
            file_type = FileType.CSV
            scraper_logger.info("File type csv")
        elif settings.SCRAPER_DOWNLOAD_FILE_TYPE == '.xlsx':
            file_type = FileType.XLSX
            scraper_logger.info("File type xlsx")
        else:
            server_logger.error('Unsupported file type in .env!')
            raise NotImplementedError

        # **Télécharger et déplacer le fichier**
        try:
            click_download_button(driver=driver, file_type=file_type)
        except Exception as e:
            scraper_logger.error(e)
            raise e
        new_file = move_and_rename_file(
            settings.SCRAPER_DOWNLOAD_ABS_PATH,
            settings.SCRAPER_FILE_DEST,
            file_type
        )

        scraper_logger.info("✅ Processus terminé avec succès !")

        message = await process_campbell_file(new_file, file_type)

        server_logger.info(message)

        return "Campbell Scraping Process terminated with success!"
    except Exception as e:
        scraper_logger.error("❌ Fermeture du Scraper avec erreur: ", e)
    finally:
        driver.quit()
        scraper_logger.info("🔒 Fermeture du navigateur.")

async def process_campbell_file(file: str, file_type: FileType, db: AsyncSession = Depends(get_db)):
    try:
        async with AsyncSessionLocal() as db:
            if file_type == FileType.XLSX:
                wb = load_workbook(file, data_only=True)
                sheet = wb.active
                server_logger.info(f"Reading File at {file} for loading to the DB.")
                # Skip the header row (start from row 2)

                row_count = 0  # Counter to track how many rows have been processed

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    timestamp_str, air_temp_avg, batt_voltage_avg, bp_mbar_avg, dew_point_avg, met_sens_status, \
                        ms60_irradiance_avg, p_temp_avg, rain_mm_tot, humidity, wind_dir, wind_speed = row

                    timestamp = datetime.strptime(timestamp_str, '%Y-%m-%dT%H:%M:%S')

                    sensor_data = CampbellMessage(
                        timestamp=timestamp.replace(tzinfo=None),
                        air_temp_avg=air_temp_avg,
                        batt_voltage_avg=batt_voltage_avg,
                        bp_mbar_avg=bp_mbar_avg,
                        dew_point_avg=dew_point_avg,
                        met_sens_status=met_sens_status,
                        ms60_irradiance_avg=ms60_irradiance_avg,
                        p_temp_avg=p_temp_avg,
                        rain_mm_tot=rain_mm_tot,
                        humidity=humidity,
                        wind_dir=wind_dir,
                        wind_speed=wind_speed
                    )

                    new_sensor = CampbellSensors(
                        timestamp=sensor_data.timestamp,
                        air_temp_avg=sensor_data.air_temp_avg,
                        batt_voltage_avg=sensor_data.batt_voltage_avg,
                        bp_mbar_avg=sensor_data.bp_mbar_avg,
                        dew_point_avg=sensor_data.dew_point_avg,
                        met_sens_status=sensor_data.met_sens_status,
                        ms60_irradiance_avg=sensor_data.ms60_irradiance_avg,
                        p_temp_avg=sensor_data.p_temp_avg,
                        rain_mm_tot=sensor_data.rain_mm_tot,
                        humidity=sensor_data.humidity,
                        wind_dir=sensor_data.wind_dir,
                        wind_speed=sensor_data.wind_speed,
                        created_at=datetime.now()  # Use the current timestamp for creation
                    )
                    db.add(new_sensor)
                    row_count += 1
                    server_logger.info(f"New Campbell sensor added to DB. Created Entity: \n{new_sensor.__dict__}")

                    if row_count % 100 == 0:  # Commit every 100 rows
                        await db.flush()  # Flush to ensure data is sent to the DB
                        await db.commit()  # Commit the transaction to the DB
                        server_logger.info(f"Committed {row_count} rows to DB.")

            if file_type == FileType.CSV:
                server_logger.info(f"Reading CSV file at {file} for loading to the DB.")

                with open(file, newline='', encoding='utf-8') as csvfile:
                    reader = csv.reader(csvfile)
                    header = next(reader)  # Skip header

                    row_count = 0  # Counter to track how many rows have been processed

                    for row in reader:
                        timestamp_str, air_temp_avg, batt_voltage_avg, bp_mbar_avg, dew_point_avg, met_sens_status, \
                            ms60_irradiance_avg, p_temp_avg, rain_mm_tot, humidity, wind_dir, wind_speed = row

                        timestamp = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M')

                        sensor_data = CampbellMessage(
                            timestamp=timestamp.replace(tzinfo=None),
                            air_temp_avg=float(air_temp_avg) if air_temp_avg else None,
                            batt_voltage_avg=float(batt_voltage_avg) if batt_voltage_avg else None,
                            bp_mbar_avg=float(bp_mbar_avg) if bp_mbar_avg else None,
                            dew_point_avg=float(dew_point_avg) if dew_point_avg else None,
                            met_sens_status=met_sens_status,
                            ms60_irradiance_avg=float(ms60_irradiance_avg) if ms60_irradiance_avg else None,
                            p_temp_avg=float(p_temp_avg) if p_temp_avg else None,
                            rain_mm_tot=float(rain_mm_tot) if rain_mm_tot else None,
                            humidity=float(humidity) if humidity else None,
                            wind_dir=float(wind_dir) if wind_dir else None,
                            wind_speed=float(wind_speed) if wind_speed else None
                        )

                        new_sensor = CampbellSensors(
                            timestamp=sensor_data.timestamp,
                            air_temp_avg=sensor_data.air_temp_avg,
                            batt_voltage_avg=sensor_data.batt_voltage_avg,
                            bp_mbar_avg=sensor_data.bp_mbar_avg,
                            dew_point_avg=sensor_data.dew_point_avg,
                            met_sens_status=sensor_data.met_sens_status,
                            ms60_irradiance_avg=sensor_data.ms60_irradiance_avg,
                            p_temp_avg=sensor_data.p_temp_avg,
                            rain_mm_tot=sensor_data.rain_mm_tot,
                            humidity=sensor_data.humidity,
                            wind_dir=sensor_data.wind_dir,
                            wind_speed=sensor_data.wind_speed,
                            created_at=datetime.now()
                        )
                        db.add(new_sensor)
                        row_count += 1
                        server_logger.info(f"New Campbell sensor added to DB. Created Entity: \n{new_sensor.__dict__}")

                        if row_count % 100 == 0:  # Commit every 100 rows
                            await db.flush()  # Flush to ensure data is sent to the DB
                            await db.commit()  # Commit the transaction to the DB
                            server_logger.info(f"Committed {row_count} rows to DB.")

                    else:
                        await db.commit()  # Commit the remaining rows after finishing processing
                        server_logger.info("Commit successful!")

    except IntegrityError as e1:
        server_logger.error(f"Error processing Campbell file: {e1}")
        server_logger.error(f"IntegrityError -- Please check if the inserted data already exists.")
        raise e1
    except Exception as e:
        # Log the error and traceback for better diagnosis
        server_logger.error("Error processing Campbell file: %s", e)
        server_logger.error("Stack trace: %s", traceback.format_exc())
        raise e

    return {"message": "File processed and saved to database."}

